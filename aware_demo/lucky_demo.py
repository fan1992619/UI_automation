import requests
import json
import time
from openpyxl import load_workbook,Workbook
'''
ID:1-->500at
ID:2-->THANKS 
ID:3-->马克杯
ID:4-->100AT
ID:5-->1000at
ID:6-->ipad
'''
def send_get(url,header,data=None):
    res=requests.post(url=url,headers=header,data=data).json()
    return res
if __name__ == '__main__':
    url='http://api.test.initialvc.com/v1/roulette/rAware'
    header_187={
        'Host': 'api.test.initialvc.com',
        'Connection': 'Keep-Alive',
        'Accept-Encoding': 'gzip',
        'User-Agent': 'okhttp/3.8.1',
        'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpdGVzdC5hdC50b3BcL3YxXC9hY2NvdW50XC9zaWduaW4iLCJpYXQiOjE1MzgwMTg0OTQsImV4cCI6MTU2OTU1NDQ5NCwibmJmIjoxNTM4MDE4NDk0LCJqdGkiOiJqTkF2TlVzVUgwMDdxSWM4Iiwic3ViIjoiMzMiLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.oGHDVe7Wb0_VvG7s2cBneH7XO6Mf2bqs8ukCDryE8ew',
        'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
        'getuiclientid': '748c2c62c0dcbc7e7c16833fec4a35f4',
        'platform': 'android',
        'userid': '33',
        'version': '2.2.1'
    }
    header_182={
        'Host': 'api.test.initialvc.com',
        'version': '2.2.1',
        'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwOlwvXC9hcGkudGVzdC5pbml0aWFsdmMuY29tXC92MVwvYWNjb3VudFwvc2lnbmluIiwiaWF0IjoxNTM4MDMwNTIxLCJleHAiOjE1Njk1NjY1MjEsIm5iZiI6MTUzODAzMDUyMSwianRpIjoiVVFpaE1PMXZOTlpKYWVXOCIsInN1YiI6IjMxIiwicHJ2IjoiYzhlZTFmYzg5ZTc3NWVjNGM3Mzg2NjdlNWJlMTdhNTkwYjZkNDBmYyJ9.dm6fpDiC4ugDuBwOyCvaD1SLCyMhvsYZZ2lGmHse7jU',
        'userid': '31',
        'Accept-Encoding': 'gzip;q=1.0, compress;q=0.5',
        'Accept-Language': 'zh-Hans-CN;q=1.0, en-CN;q=0.9',
        'platform': 'iOS',
        'Accept': '*/*',
        'deviceid': 'd05455f5bd20437199c778aa46b7618e',
        'User-Agent': 'AWARE-iOS/2.2.1 (top.at.aware; build:003; iOS 11.4.0) Alamofire/4.7.2',
        'Connection': 'keep-alive',
        'Getuiclientid': '654e8dae6f761ebe0ae73c3d74615deb'
    }
    for i in range(90):
        wb = load_workbook('../aware_demo/lucky.xlsx')
        ws = wb.active
        ws.title = "抽奖记录"
        res=send_get(url,header_187)
        # print(res)
        lucky_id=res['data']['jpId']
        lucky_id_num=int(lucky_id)
        # content_lucky=("第%d次奖励id为{0}--->500AT".format(lucky_id_num) % i)
        try:
            if lucky_id_num==20:
                content_lucky = ("第%d次奖励id为{0}--->以太坊卡".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励id为{0}--->以太坊卡".format(lucky_id_num) % i,res)
            elif lucky_id_num==21:
                content_lucky = ("第%d次奖励id为{0}--->thanks".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励id为{0}--->thanks".format(lucky_id_num) % i, res)
            elif lucky_id_num==22:
                content_lucky = ("第%d次奖励id为{0}--->P2 plus钱包".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励id为{0}--->P2 plus钱包".format(lucky_id_num) % i, res)
            elif lucky_id_num==23:
                content_lucky = ("第%d次奖励id为{0}--->库神比特币卡片".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励为{0}--->库神比特币卡片".format(lucky_id_num) % i, res)
            elif lucky_id_num==24:
                content_lucky = ("第%d次奖励id为{0}--->白皮书".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励id为{0}--->白皮书".format(lucky_id_num) % i, res)
            elif lucky_id_num ==25:
                content_lucky = ("第%d次奖励id为{0}--->库神钱包优惠码".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次奖励为id{0}--->库神钱包优惠码".format(lucky_id_num) % i, res)
            else:
                content_lucky = ("第%d次奖励id为{0}--->抽奖次数用完必".format(lucky_id_num) % i)
                ws['A{0}'.format(i + 1)] = content_lucky
                print("第%d次抽奖次数用完必" % i,res)
        except:
            pass
        finally:
            wb.save('../aware_demo/lucky.xlsx')
        # time.sleep(0.5)
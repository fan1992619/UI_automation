#coding=utf-8
#18210542401 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM
#18782610762 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw
from lxml import html
from openpyxl import load_workbook
import requests
import random
import xlrd
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import data_config
import json
str_182='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM'
str_187='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw'
class SendArticles():
    #初始化表单，添加到excel表中
    def __init__(self):
        self.wb = load_workbook('../aware_demo/title.xlsx')
        self.ws = self.wb.active
        self.ws.title = "抓取标题"
        self.ws.sheet_properties.tabColor = 'ff0000'
        url='https://www.jinse.com/news/bitcoin'
        res=requests.get(url=url)
        res.encoding='utf-8'
        html_doc=res.text
        #获取xpath对象
        selector=html.fromstring(html_doc)
        #找到列表集合
        ui_list=selector.xpath('//div[@class="wrap margin-b clearfix"]/div[@class="wrap-left left"]/div/div[@class="content"]/div[@class="list clearfix"]')
        # print(len(ui_list))
        # title=ui_list[19].xpath('//div[@class="post right"]/div/a/@title')
        # link=ui_list[19].xpath('//div[@class="post right"]/div/a/@href')
        # print(title[19],link[19])
        for i,li in enumerate(ui_list):
            title=li.xpath('//div[@class="post right"]/div/a/@title')
            link=li.xpath('//div[@class="post right"]/div/a/@href')
            # print('第%d次' % i,title[i],link[i])
            self.ws['A{0}'.format(i+2)]=title[i]
            self.ws['B{0}'.format(i+2)]=link[i]
            i+=1
        self.wb.save('../aware_demo/title.xlsx')
    #获取url
    def open_url(self,row):
        data = xlrd.open_workbook('../aware_demo/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_url())
        url =tables.cell_value(row, col)
        return url
    #遍历url
    def articles_url(self):
        data = xlrd.open_workbook('../aware_demo/title.xlsx')
        tables_num = data.sheets()[0].nrows
        for i in range(tables_num):
            if i<1:
                continue
            return self.open_url(i)
    #获取title
    def open_title(self,row):
        data = xlrd.open_workbook('../aware_demo/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_title())
        title =tables.cell_value(row, col)
        return title
    #发送文章
    def send_atricl(self):
        data = xlrd.open_workbook('../aware_demo/title.xlsx')
        tables = data.sheets()[0].nrows
        print(tables)
        for i in range(tables):
            if i<1:
                continue
            link=self.open_url(i)
            title=self.open_title(i)
            url='https://api.at.top/v1/projects/22/articles'
            data={
                'type': '1',
                'link': '{0}'.format(link),
                'title':'{0}'.format(title)
                # 'timestamp': '1537583500',
                # 'signature': '43fa84601a94e0b5920b5ce7a73c680f'
            }
            header_at={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Content-Length': '764',
                'Host': 'api.at.top',
                'Connection': 'Keep-Alive',
                'Accept-Encoding': 'gzip',
                'User-Agent': 'okhttp/3.8.1',
                'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwOlwvXC9hcGkudGVzdC5pbml0aWFsdmMuY29tXC92MVwvYWNjb3VudFwvc2lnbmluIiwiaWF0IjoxNTM3NTMxMDk3LCJleHAiOjE1NjkwNjcwOTcsIm5iZiI6MTUzNzUzMTA5NywianRpIjoiSk9VaUN2QVZBMVNGa05iZyIsInN1YiI6IjMzIiwicHJ2IjoiYzhlZTFmYzg5ZTc3NWVjNGM3Mzg2NjdlNWJlMTdhNTkwYjZkNDBmYyJ9.FExRUmpERoEDDUtFa666lxHjQhbBeD4TQKtH_sa6Jrw',
                'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
                'getuiclientid': '330ccf5988efd42e629f88e533488d4e',
                'platform': 'android',
                'userid': '33',
                'version': '2.2.0'
            }
            # print('第%d次' % i, res)
            header={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Content-Length': '239',
                'Connection': 'Keep-Alive',
                'Accept-Encoding': 'gzip',
                'User-Agent': 'okhttp/3.8.1',
                'Host': 'api.at.top',
                'Authorization': str_182,
                'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
                'platform': 'android',
                'userid': '',
                'version': '1.2.0'
            }
            res=requests.post(url=url,data=data,headers=header).json()
            print ('第%d次' % i,res)
if __name__ == '__main__':
    send = SendArticles()
    send.articles_url()
#coding=utf-8
#18210542401 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM
#18782610762 bearer     Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw
from lxml import html
# import xlsxwriter
# import urllib.request
from openpyxl import load_workbook,Workbook
import requests
import urllib.request
import random
import xlrd
from openpyxl.drawing.image import Image
from requests.packages.urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import data_config
import json
str_182='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NTIyNCwiZXhwIjoxNTY1NDIxMjI0LCJuYmYiOjE1MzM4ODUyMjQsImp0aSI6Im80ZmFKVnhlMUxXSmlmdksiLCJzdWIiOjMxLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.UCbB_ilaw1xu92aB4oQ5k_dJXT6tj-xHVGtO5-NiHTM'
str_187='Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczpcL1wvYXBpLmF0LnRvcFwvdjFcL2FjY291bnRcL3NpZ25pbiIsImlhdCI6MTUzMzg4NDY0NywiZXhwIjoxNTY1NDIwNjQ3LCJuYmYiOjE1MzM4ODQ2NDcsImp0aSI6ImlGTlJkTTRSVk90S3ZhWTgiLCJzdWIiOjMzLCJwcnYiOiJjOGVlMWZjODllNzc1ZWM0YzczODY2N2U1YmUxN2E1OTBiNmQ0MGZjIn0.uMkAI8VR6lZOCr27znRYLfkZRazvpvxDHjc8wBi1xPw'
class SendArticles():
    def spider_article(self):
        wb = load_workbook('../appium_android/title.xlsx')
        ws = wb.active
        ws.title = "抓取标题"
        ws.sheet_properties.tabColor = 'ff0000'
        url = 'https://www.jinse.com/news/bitcoin'
        res = requests.get(url=url, verify=False)
        res.encoding = 'utf-8'
        html_doc = res.text
        # 获取xpath对象
        selector = html.fromstring(html_doc)
        # 找到列表集合
        ui_list = selector.xpath(
            '//div[@class="wrap margin-b clearfix"]/div[@class="wrap-left left"]/div/div[@class="content"]/div[@class="list clearfix"]')
        for i, li in enumerate(ui_list):
            title = li.xpath('//div[@class="post right"]/div/a/@title')
            link = li.xpath('//div[@class="post right"]/div/a/@href')
            content=li.xpath('//div[@class="post right"]/div[@class="message"]/text()')
            img_url=li.xpath('//div[@class="image left"]/a/img/@src')
            # print(content[i])
            # print('第%d次' % i,title[i],link[i],content[i],img_url[i])
            ws['A{0}'.format(i + 2)] = title[i]
            ws['B{0}'.format(i + 2)] = link[i]
            ws['C{0}'.format(i + 2)]=content[i]
            ws['D{0}'.format(i + 2)]=img_url[i]
            i += 1
        wb.save('../appium_android/title.xlsx')
    #获取url
    def open_url(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_url())
        url =tables.cell_value(row, col)
        # print(url)
        return url
    #获取title
    def open_title(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_title())
        title =tables.cell_value(row, col)
        # print(title)
        return title
    #获取文章内容
    def open_content(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_content())
        content =tables.cell_value(row, col)
        # print(title)
        return content
    #获取文章图片的url
    def open_img_url(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_img_url())
        img_url =tables.cell_value(row, col)
        # print(title)
        return img_url
    #获取这张图片内容
    def open_img_content(self,row):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        col = int(data_config.get_img_content())
        img_content =tables.cell_value(row, col)
        # print(title)
        return img_content
    #遍历url
    def articles_url(self):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables_num = data.sheets()[0].nrows
        for i in range(tables_num):
            if i<1:
                continue
            return self.open_url(i)
    #爬取图片储存并放入到excel表中
    def spider_img(self):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0]
        tables=tables.nrows
        tables=random.randint(0,tables)
        return tables
    #插入图片暂时不能允许
    '''
    def spider_images(self):
        self.spider_article()
        data = load_workbook('../appium_android/title.xlsx')
        ws = data.active
        # data.sheetnames('抓取标题'）
        # tables_num = data.sheets()[0].nrows
        for i in range(19):
            if i<1:
                continue
            else:
                img_url=self.open_img_url(i)
                get_img=self.spider_img()
                print(get_img)
                # print(img_url,)
                # f =open('../appium_android/title.xlsx',"wb")# 打开文件
                req=urllib.request.urlopen(get_img)
                buf=req.red()# 读出文件
                f.write(buf)# 写入文件
                img=Image('../appium_android/{0}.jpg'.format(i))
                # f.writelines(ws['D{0}'.format(i + 2)])
                # img=Image('../appium_android/{0}.jpg'.format(i))
                ws.add_image(img,'E{0}'.format(i+2))
                i+=1
            data.save('../appium_android/title.xlsx')
    '''
    #测试插入图片
    def img_test():
        listurl = ['https://contestimg.wish.com/api/webimage/59647c7e7baa287c180fa0e0-3-original.jpg',
                   'https://contestimg.wish.com/api/webimage/59647c7e7baa287c180fa0e0-original.jpg', ]

        # 根据图片链接列表，把图片保存到本地
        i = 0
        for url in listurl:
            f = open(str(i) + '.jpg', "wb")  # 打开文件
            req = urllib.request.urlopen(url)
            buf = req.read()  # 读出文件
            f.write(buf)  # 写入文件
            i = i + 1
        # 将图片一次导入到表格的1，2...行
        data = load_workbook('../appium_android/pict.xlsx')
        ws = data.active
        img = Image('../appium_android/0.jpg')
        ws.add_image(img, 'E2')
        data.save('../appium_android/title.xlsx')
    #发送文章
    def send_atricl(self):
        self.spider_article()
        data = xlrd.open_workbook('../appium_android/title.xlsx')
        tables = data.sheets()[0].nrows
        print(tables)
        for i in range(tables):
            if i<1:
                continue
            link=self.open_url(i)
            title=self.open_title(i)
            url='https://api.at.top/v1/projects/22/articles'
            data={
                'type': '1',
                'link': '{0}'.format(link),
                'title':'{0}'.format(title)
                # 'timestamp': '1537583500',
                # 'signature': '43fa84601a94e0b5920b5ce7a73c680f'
            }
            header_at={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Content-Length': '764',
                'Host': 'api.at.top',
                'Connection': 'Keep-Alive',
                'Accept-Encoding': 'gzip',
                'User-Agent': 'okhttp/3.8.1',
                'Authorization': 'Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwOlwvXC9hcGkudGVzdC5pbml0aWFsdmMuY29tXC92MVwvYWNjb3VudFwvc2lnbmluIiwiaWF0IjoxNTM3NTMxMDk3LCJleHAiOjE1NjkwNjcwOTcsIm5iZiI6MTUzNzUzMTA5NywianRpIjoiSk9VaUN2QVZBMVNGa05iZyIsInN1YiI6IjMzIiwicHJ2IjoiYzhlZTFmYzg5ZTc3NWVjNGM3Mzg2NjdlNWJlMTdhNTkwYjZkNDBmYyJ9.FExRUmpERoEDDUtFa666lxHjQhbBeD4TQKtH_sa6Jrw',
                'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
                'getuiclientid': '330ccf5988efd42e629f88e533488d4e',
                'platform': 'android',
                'userid': '33',
                'version': '2.2.0'
            }
            # print('第%d次' % i, res)
            header={
                'Content-Type': 'application/x-www-form-urlencoded',
                'Content-Length': '239',
                'Connection': 'Keep-Alive',
                'Accept-Encoding': 'gzip',
                'User-Agent': 'okhttp/3.8.1',
                'Host': 'api.at.top',
                'Authorization': str_182,
                'deviceid': 'ac:c1:ee:c0:33:34-ac:c1:ee:c0:33:34',
                'platform': 'android',
                'userid': '',
                'version': '1.2.0'
            }
            res=requests.post(url=url,data=data,headers=header,verify=False).json()
            print ('第%d次' % i,res)
if __name__ == '__main__':
    send = SendArticles()
    # send.open_url(3)
    # send.open_title(3)
    send.spider_article()
    # send.spider_images()
    # print(send.articles_url())
    # print(send.open_img_url(3))
    # print(send.open_content(3))
    # send.spider_img()